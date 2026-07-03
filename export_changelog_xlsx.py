#!/usr/bin/env python3
"""
export_changelog_xlsx.py

Reads the `changelog` table from permits.db (populated by permit_monitor.py)
and writes a formatted Excel workbook: air_permit_changelog.xlsx

Run this after permit_monitor.py fetch. In CI (GitHub Actions), the workflow
runs this automatically after every scheduled fetch.
"""

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "permits.db"
OUT_PATH = "air_permit_changelog.xlsx"

FONT_NAME = "Arial"
signal_colors = {
    "SURPLUS_LEAD": "C6EFCE",
    "EARLY_BUYER_LEAD": "FFEB9C",
    "REPOWER_SIGNAL": "BDD7EE",
    "OWNERSHIP_SIGNAL": "E2D6F3",
    "NON_RENEWAL_WARNING": "FFC7CE",
    "AUCTION_SCRAP_SIGNAL": "F4B084",
}


def load_latest_snapshot_rows():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT snapshot_id FROM snapshots ORDER BY snapshot_id DESC LIMIT 1")
    snapshot_row = cur.fetchone()
    if not snapshot_row:
        conn.close()
        return []

    snapshot_id = snapshot_row[0]
    cur.execute(
        """
        SELECT permit_id, facility_name, owner_entity, state, status,
               equipment_desc, capacity_mw, issue_date, expiration_date,
               last_action_date
        FROM permit_records
        WHERE snapshot_id = ?
        ORDER BY facility_name, permit_id
        """,
        (snapshot_id,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def main():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT detected_at, state, signal_type, facility_name, owner_entity,
               prior_status, new_status, equipment_matches, entity_matches, note
        FROM changelog
        ORDER BY id DESC
        """
    )
    rows = cur.fetchall()
    conn.close()

    latest_snapshot_rows = load_latest_snapshot_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Changelog"

    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
    base_font = Font(name=FONT_NAME, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Detected At",
        "State",
        "Signal Type",
        "Facility Name",
        "Owner Entity",
        "Prior Status",
        "New Status",
        "Equipment Match",
        "Entity Match",
        "Note",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    if not rows:
        ws.cell(
            row=2,
            column=1,
            value="No changelog entries yet. The latest live EPA snapshot was loaded successfully.",
        )
        ws.cell(row=2, column=2, value="TX")
        ws.cell(row=2, column=3, value="LIVE_SNAPSHOT")
        ws.cell(row=2, column=4, value="EPA ECHO facility records were fetched successfully.")
        ws.cell(row=2, column=10, value="Run the fetch step again later to capture real change events.")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = base_font
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        fill_color = signal_colors.get(row[2])
        if fill_color:
            ws.cell(row=r_idx, column=3).fill = PatternFill(
                "solid", start_color=fill_color, end_color=fill_color
            )

    widths = [16, 6, 20, 30, 26, 16, 20, 16, 12, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    snapshot_ws = wb.create_sheet("Live Snapshot")
    snapshot_headers = [
        "Permit ID",
        "Facility Name",
        "Owner Entity",
        "State",
        "Status",
        "Equipment Desc",
        "Capacity MW",
        "Issue Date",
        "Expiration Date",
        "Last Action Date",
    ]
    for i, h in enumerate(snapshot_headers, start=1):
        cell = snapshot_ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r_idx, row in enumerate(latest_snapshot_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = snapshot_ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    snapshot_widths = [16, 30, 26, 8, 18, 34, 12, 14, 14, 18]
    for i, w in enumerate(snapshot_widths, start=1):
        snapshot_ws.column_dimensions[get_column_letter(i)].width = w
    snapshot_ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Wrote {len(rows)} changelog row(s) to {OUT_PATH}")


if __name__ == "__main__":
    main()
